from django.shortcuts import render,redirect



from .models import Cliente
from .forms import ClienteForm
from .models import Ventas
from .forms import VentasForm
# LOGIN
from django.shortcuts import render
from django.contrib import messages




#EXPORTACION EXCEL
import openpyxl
from django.http import HttpResponse
from django.views import View
from .models import Cliente
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
import datetime


#GRAFICAR
import matplotlib.pyplot as plt




def inicio(request):
     return render(request, 'paginas/inicio.html')
 
 
 
 # CRUD CLIENTES -----------------

def clientes(request):
    clientes = Cliente.objects.all()
    return render(request, 'clientes/index.html',{'clientes': clientes})




def crear(request):
    formulario= ClienteForm(request.POST or None)
    if formulario.is_valid():
        formulario.save()
        return redirect('clientes')
    return render(request, 'clientes/crear.html', {'formulario': formulario})


def editar(request,id):
    cliente= Cliente.objects.get(id=id)
    formulario= ClienteForm(request.POST or None,instance=cliente)
    if  formulario.is_valid() and request.POST:
        formulario.save()
        return redirect('clientes')
    return render(request, 'clientes/editar.html',{'formulario': formulario})


def eliminar(request, id):
    cliente =  Cliente.objects.get(id=id)
    cliente.delete()
    return redirect('clientes')


 # CRUD CLIENTES -----------------
 
 
 
 
 
 #CRUD VENTAS -------------------
 
#def ventas(request):
 #   return render(request,'ventas/crearV.html') 
 
 
def ventas(request):
    ventas = Ventas.objects.all()
    
    return render(request, 'ventas/indexV.html',{'ventas': ventas})

 
def crear_ventas(request):
    formularioV= VentasForm(request.POST or None)
    if formularioV.is_valid():
        formularioV.save()
        return redirect('ventas')
    return render(request,'ventas/crearV.html', {'formularioV': formularioV})
 
def editar_ventas(request,id):
    ventas= Ventas.objects.get(id=id)
    formularioV= VentasForm(request.POST or None, instance=ventas)
    if formularioV.is_valid() and request.POST:
        formularioV.save();
        return redirect('ventas')
    return render(request,'ventas/editarV.html',{'formularioV': formularioV})








def eliminar_ventas(request, id):
    ventas =  Ventas.objects.get(id=id)
    ventas.delete()
    return redirect('ventas')


 #CRUD VENTAS ----------------
 
 
 
 
 
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']

        if username == 'hola' and password == 'hola':
            # Iniciar sesión exitoso
            # Puedes redirigir a una página de inicio o mostrar un mensaje de éxito
             return render(request, 'paginas/inicio.html')

        else:
            # Credenciales inválidas
            messages.error(request, 'Credenciales inválidas')

    return render(request, 'login.html')
 
 
 
 
 
 
 ##Expotar a Excel ----------------------------------------------------------
def exportarexcel(request):
     return render(request, 'paginas/exportarexcel.html')
 
def export_data_to_excel(request):
    # Obtén los datos de la base de datos
    data = Cliente.objects.all()  # Reemplaza "TuModelo" con el nombre de tu modelo de base de datos

    # Crea un nuevo archivo de Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Define estilos de fuente
    header_font = Font(name='Arial', bold=True, size=12)
    data_font = Font(name='Arial', size=11)

    # Establece el formato de celda para fechas
    date_format = NamedStyle(name='date_format')
    date_format.number_format = 'dd-mm-yyyy'
    workbook.add_named_style(date_format)

    # Establece los estilos de alineación
    header_alignment = Alignment(horizontal='center', vertical='center')
    data_alignment = Alignment(horizontal='center', vertical='center')

    # Establece los estilos de relleno y borde
    header_fill = PatternFill(start_color='E9E9E9', end_color='E9E9E9', fill_type='solid')
    header_border = Border(bottom=Side(border_style='thin'))
    field_border = Border(
        left=Side(border_style='thin', color='00FF00'),
        right=Side(border_style='thin', color='00FF00'),
        top=Side(border_style='thin', color='00FF00'),
        bottom=Side(border_style='thin', color='00FF00')
    )
    field_fill = PatternFill(start_color='C1FFC1', end_color='C1FFC1', fill_type='solid')

    # Agrega los encabezados de las columnas
    headers = ['id', 'Nombre', 'Apellido_paterno','Apellido_materno','edad','celular','genero' ]  # Reemplaza con los nombres de tus campos

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = header_border

    # Agrega los datos a las filas
    for row_num, item in enumerate(data, 2):
        row = [item.id, item.nombre, item.apelidopate, item.apelidomater,item.edad,item.celular,item.genero]  # Reemplaza con los nombres de tus campos

        for col_num, value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num, value=value)

            if col_num == 1:  # Ajusta para el campo ID
                cell.number_format = '0'  # Formato de número entero sin decimales
                cell.alignment = data_alignment
            elif isinstance(value, str):
                cell.font = data_font
                cell.alignment = data_alignment
            elif isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.font = data_font
                cell.alignment = data_alignment
            elif isinstance(value, datetime.date):
                cell.style = 'date_format'
                cell.alignment = data_alignment

            cell.border = field_border
            cell.fill = field_fill

    # Ajusta el ancho de las columnas y establece el autoajuste
    for column_cells in worksheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value)) for cell in column_cells)
        adjusted_width = max_length + 4  # Aumenta el ancho para una mejor visualización
        worksheet.column_dimensions[column_letter].width = adjusted_width
        worksheet.column_dimensions[column_letter].bestFit = True

    # Configura la respuesta HTTP para descargar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=datos.xlsx'

    # Guarda el archivo Excel en la respuesta HTTP
    workbook.save(response)

    return response
 
 #GRAFICAR
def graficas(request):
     return render(request, 'paginas/graficas.html')
 
def grafico(request):
    # Datos para el gráfico
    x = [1, 2, 3, 4, 5]
    y = [10, 20, 15, 25, 30]
    
    # Generar gráfico
    plt.plot(x, y)
    plt.xlabel('Eje X')
    plt.ylabel('Eje Y')
    plt.title('Mi gráfico')
    
    # Guardar gráfico en un archivo
    plt.savefig('/projectgravitii/imagenes/grafico.png')
    
    return HttpResponse("Gráfico generado")
 
 
    
 
# Create your views here.
